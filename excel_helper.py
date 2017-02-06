# -*- coding: utf-8 -*-

import logging
import datetime

from enum import Enum
from openpyxl import load_workbook
from calendar import TimeEncoding, month_name
from datetime import date as sysdate

logger = logging.getLogger(__name__)


class ExcelHelper:

    NAME_CELL = 'C3'
    NAME = "Rüegg Simon"
    PENSUM_CELL = 'I5'
    PENSUM = .80  # 80%
    MORNING_IN_COL = 3
    MORNING_OUT_COL = 4
    AFTERNOON_IN_COL = 6
    AFTERNOON_OUT_COL = 7

    def __init__(self, excel_file):
        self.workbook = load_workbook(filename=excel_file, data_only=False)
        self.sheet = None
        self.cell = None

    def write_entry(self, date, time, type):
        sheet = self.find_sheet(date)

        sheet = self.sheet[1]
        sheet[self.NAME_CELL] = self.NAME
        sheet[self.PENSUM_CELL] = self.PENSUM

        self.find_cell(date)
        cell = self.cell
        col = None
        time_value = datetime.datetime.strptime(time, '%H:%M').time()
        time_dec = time_value.hour + round(time_value.minute / float(60), 2)
        if type == EntryType.MORNING_IN.value:
            col = self.MORNING_IN_COL
        elif type == EntryType.MORNING_OUT.value:
            col = self.MORNING_OUT_COL
        elif type == EntryType.AFTERNOON_IN.value:
            col = self.AFTERNOON_IN_COL
        elif type == EntryType.AFTERNOON_OUT.value:
            col = self.AFTERNOON_OUT_COL
        else:
            logger.error("Unknown entry type '%d' on %s!", type, date)

        if not col is None:
            sheet.cell(row=cell.row, column=col, value=time_dec)
            logger.debug("Wrote time '%.2f' to field '%s'", time_dec, cell)

    def find_sheet(self, date):
        if not self.sheet is None and self.sheet[0].month == date.month:
            return self.sheet
        else:
            name = self.get_month_name(date.month, "de_CH.utf8")
            y = str(date.year % 100)
            logger.debug("Search sheet with name '%s' and year '%s'", name, y)
            for sheet_name in self.workbook.get_sheet_names():
                if sheet_name.startswith(name) and sheet_name.endswith(y):
                    logger.debug("Found sheet '%s'", sheet_name)
                    self.sheet = (date, self.workbook[sheet_name])
                    return self.sheet

    def find_cell(self, date):
        if not self.cell is None and date == self.cell.value.date():
            return self.cell
        else:
            try:
                logger.debug("Search cell with date '%s'", date)
                for row in self.sheet[1].iter_rows(min_row=11, max_row=50, min_col=2, max_col=2):
                    for cell in row:
                        if not cell.value is None:
                            if cell.value.date() == date:
                                logger.debug("Found cell (%s) for date (%s)", cell, date)
                                self.cell = cell
                                return self.cell
            except:
                print cell

    def get_month_name(self, month_no, locale):
        with TimeEncoding(locale) as encoding:
            name = month_name[month_no]
            return name

    def save(self, path):
        self.workbook.save(path)


class EntryType(Enum):
    MORNING_IN = 0
    MORNING_OUT = 1
    AFTERNOON_IN = 2
    AFTERNOON_OUT = 3
